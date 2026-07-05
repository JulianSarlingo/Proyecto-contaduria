import pandas as pd
import re
import json
import os
import time
from openpyxl import load_workbook
from os import system

system('cls')


# ==== helpers generales ====

def get_file_mtime(path: str) -> float:
    """Devuelve la fecha de última modificación de un archivo."""
    return os.path.getmtime(path)


# ==== helpers para contar filas ====

def contar_filas_utiles(ruta_excel, start_row=9, col_clave="A"):
    """
    Cuenta cuántas filas con datos reales hay a partir de start_row.
    Usa la columna clave (por defecto A) para saber si la fila está vacía.
    """
    wb = load_workbook(ruta_excel, read_only=True, data_only=True)
    ws = wb.active

    fila_actual = start_row
    while True:
        valor = ws[f"{col_clave}{fila_actual}"].value
        if valor is None or str(valor).strip() == "":
            break
        fila_actual += 1

    return fila_actual - start_row  # cantidad de filas con datos


# ==== paso 1: leer Excel y devolver DataFrame ====

def configuracion(ruta_excel, sheet_name):
    """
    Lee la configuración desde un Excel, detectando automáticamente la cantidad de filas útiles.
    Devuelve un DataFrame limpio con solo las columnas necesarias.
    """
    """Columnas 1, 8, 172, 4, 21, 7, respectivamente:"""
    columnas_utiles = ["Clientes", "CUIT", "Clave Afip", "Contribuyente", "e-mail", "tipo monotributo"]

    n_filas = contar_filas_utiles(ruta_excel, start_row=10, col_clave="A")

    df = pd.read_excel(
        ruta_excel,
        sheet_name=sheet_name,
        skiprows=8,  # encabezado en fila 9
        nrows=n_filas,
        usecols=columnas_utiles,
        engine="openpyxl"
    ).fillna("")

    # Normalizar encabezados (quita espacios)
    df.rename(columns=lambda c: str(c).strip(), inplace=True)

    return df


# ==== helpers para CUIT ====

def _digits11(s: str) -> str:
    """Devuelve solo los dígitos (11) de un CUIT o '' si no cumple."""
    try:
        s = int(s)
    except ValueError:
        return ""
    d = re.sub(r"\D", "", str(s))
    return d if len(d) == 11 else ""


def _es_persona(cuit: str) -> bool:
    """Persona = CUIT que NO empieza por 30."""
    d = _digits11(cuit)
    return bool(d) and not d.startswith("30")


# ==== paso 2: indexar sociedades ====

def indexar_sociedades(df):
    sociedades_por_dueno = {}
    for _, r in df.iterrows():
        cuit_fila  = _digits11(r["CUIT"])
        dueno_cuit = _digits11(r["tipo monotributo"])

        if cuit_fila.startswith("30") and dueno_cuit:
            sociedades_por_dueno.setdefault(dueno_cuit, []).append({
                "nombre": str(r["Clientes"]).strip(),
                "cuit": cuit_fila
            })
    return sociedades_por_dueno


# ==== paso 3: construir config final ====

def construir_config(df, sociedades_por_dueno):
    config = []
    vistos = set()

    for _, row in df.iterrows():
        nombre   = str(row["Clientes"]).strip()
        password = str(row["Clave Afip"]).strip()
        email    = str(row["e-mail"]).strip()
        tipo     = str(row["Contribuyente"]).strip()
        cuit_p   = _digits11(row["CUIT"])

        if not _es_persona(cuit_p):
            continue
        if cuit_p in vistos:
            continue

        config.append({
            "nombre": nombre,
            "cuit": cuit_p,
            "password": password,
            "tipo_monotributo": tipo,
            "email": email,
            "sociedades": sociedades_por_dueno.get(cuit_p, [])
        })
        vistos.add(cuit_p)

    return config


# ==== caché JSON ====

def _ruta_cache(ruta_excel: str) -> str:
    """Devuelve la ruta del archivo de caché (junto al Excel)."""
    carpeta = os.path.dirname(ruta_excel)
    return os.path.join(carpeta, ".cache_config.json")


# ==== función orquestadora ====

def procesar_config(ruta_excel, sheet_name, forzar=False):
    """
    Orquesta todo: lee el Excel (o el caché si está disponible),
    indexa sociedades y arma config final.

    La primera vez que se ejecuta, parsea el Excel y guarda una base de datos
    local (.cache_config.json) junto al archivo Excel.
    En ejecuciones posteriores, si el Excel no fue modificado y la hoja no
    cambió, carga los datos desde el caché (mucho más rápido).
    Para forzar un re-escaneo, simplemente guardá el Excel y volvé a ejecutar.
    """
    cache_file = _ruta_cache(ruta_excel)
    mtime      = get_file_mtime(ruta_excel)

    # --- Intentar cargar desde caché (salvo re-lectura forzada) ---
    if forzar:
        print("[Caché] Re-lectura forzada: ignorando el caché y releyendo el Excel.")
    if not forzar and os.path.exists(cache_file):
        try:
            with open(cache_file, 'r', encoding='utf-8') as f:
                cache_data = json.load(f)
            # El caché es válido si el Excel no cambió y la hoja coincide
            if (cache_data.get("mtime") == mtime and
                    cache_data.get("sheet_name") == sheet_name):
                n = len(cache_data["config"])
                print(f"[Caché] {n} usuarios cargados desde caché local "
                      f"(sin leer el Excel).")
                return cache_data["config"]
            else:
                print("[Caché] El Excel fue modificado o la hoja cambió. "
                      "Re-escaneando...")
        except Exception as e:
            print(f"[Caché] No se pudo leer el caché: {e}. Re-escaneando...")

    # --- Leer Excel (primera vez o si el Excel cambió) ---
    nombre_excel = ruta_excel.split('\\')[-1]
    print(f"Leyendo excel en: {nombre_excel}")
    inicio     = time.perf_counter()
    df         = configuracion(ruta_excel, sheet_name)
    sociedades = indexar_sociedades(df)
    config     = construir_config(df, sociedades)
    fin        = time.perf_counter()
    print(f"Configuración construida correctamente en {fin - inicio:.2f} segundos "
          f"({len(config)} usuarios)")

    # --- Guardar caché para la próxima ejecución ---
    try:
        with open(cache_file, 'w', encoding='utf-8') as f:
            json.dump(
                {"mtime": mtime, "sheet_name": sheet_name, "config": config},
                f,
                ensure_ascii=False,
                indent=2
            )
        print(f"[Caché] Base de datos guardada en: {os.path.basename(cache_file)}")
    except Exception as e:
        print(f"[Caché] No se pudo guardar el caché: {e}")

    return config

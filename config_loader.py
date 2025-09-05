import pandas as pd
import re
from openpyxl import load_workbook
from os import system
system('cls')



import pandas as pd
import re
from openpyxl import load_workbook

# ==== helpers ====
def contar_filas_utiles(ruta_excel, start_row=9, col_clave="A"):
    """
    Cuenta cuántas filas con datos reales hay a partir de start_row.
    Usa la columna clave (por defecto A) para saber si la fila está vacía.
    """
    wb = load_workbook(ruta_excel, read_only=True, data_only=True)
    ws = wb.active

    fila_actual = start_row
    while True:
        valor = ws[f"{col_clave}{fila_actual}"].value
        if valor is None or str(valor).strip() == "":
            break
        fila_actual += 1

    return fila_actual - start_row  # cantidad de filas con datos


# ==== paso 1: leer Excel y devolver DataFrame ====
def configuracion(ruta_excel):
    """
    Lee la configuración desde un Excel, detectando automáticamente la cantidad de filas útiles.
    Devuelve un DataFrame limpio con solo las columnas necesarias.
    """
    columnas_utiles = ["Clientes", "cuit ", "Clave Afip", "e-mail", "tipo monotributo"]

    n_filas = contar_filas_utiles(ruta_excel, start_row=10, col_clave="A")

    df = pd.read_excel(
        ruta_excel,
        skiprows=8,  # encabezado en fila 9
        nrows=n_filas,
        usecols=columnas_utiles,
        engine="openpyxl"
    ).fillna("")

    # Normalizar encabezados (quita espacios)
    df.rename(columns=lambda c: str(c).strip(), inplace=True)

    return df


# ==== helpers para CUIT ====
def _digits11(s: str) -> str:
    """Devuelve solo los dígitos (11) de un CUIT o '' si no cumple."""
    d = re.sub(r"\D", "", str(s))
    return d if len(d) == 11 else ""

def _es_persona(cuit: str) -> bool:
    """Persona = CUIT que NO empieza por 30."""
    d = _digits11(cuit)
    return bool(d) and not d.startswith("30")


# ==== paso 2: indexar sociedades ====
def indexar_sociedades(df):
    sociedades_por_dueno = {}
    for _, r in df.iterrows():
        cuit_fila  = _digits11(r["cuit"])
        dueno_cuit = _digits11(r["tipo monotributo"])

        if cuit_fila.startswith("30") and dueno_cuit:
            sociedades_por_dueno.setdefault(dueno_cuit, []).append({
                "nombre": str(r["Clientes"]).strip(),
                "cuit": cuit_fila
            })
    return sociedades_por_dueno


# ==== paso 3: construir config final ====
def construir_config(df, sociedades_por_dueno):
    config = []
    vistos = set()

    for _, row in df.iterrows():
        nombre   = str(row["Clientes"]).strip()
        password = str(row["Clave Afip"]).strip()
        email    = str(row["e-mail"]).strip()
        cuit_p   = _digits11(row["cuit"])

        if not _es_persona(cuit_p):
            continue
        if cuit_p in vistos:
            continue

        config.append({
            "nombre": nombre,
            "cuit": cuit_p,
            "password": password,
            "email": email,
            "sociedades": sociedades_por_dueno.get(cuit_p, [])
        })
        vistos.add(cuit_p)

    return config


# ==== función orquestadora ====
def procesar_config(ruta_excel):
    """
    Orquesta todo: lee el Excel, indexa sociedades, arma config final.
    """
    df = configuracion(ruta_excel)
    sociedades = indexar_sociedades(df)
    return construir_config(df, sociedades)

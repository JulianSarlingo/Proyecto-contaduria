import pandas as pd
import re
from openpyxl import load_workbook
from os import system
system('cls')

# Funcion obsoleta
def cargar_configuracion(ruta_excel):
    """
    Lee el archivo Excel de configuración y devuelve una lista de dicts con los datos por persona.

    Args:
        ruta_excel (str): Ruta al archivo Excel.

    Returns:
        list[dict]: Lista de personas con sus configuraciones.
    """
    # df = pd.read_excel(ruta_excel).fillna("")
    df = pd.read_excel(ruta_excel, skiprows=8, engine="openpyxl").fillna("")

    configuraciones = []

    for _, row in df.iterrows():
        nombre = row["Nombre"].strip()
        cuit = str(row["CUIT"]).strip()
        password = row["Contraseña AFIP"].strip()

        sociedades = []
        if row.get("Sociedad1 Check", "").strip().upper() == "X" and row.get("Sociedad1"):
            sociedades.append(row["Sociedad1"].strip())
        if row.get("Sociedad2 Check", "").strip().upper() == "X" and row.get("Sociedad2"):
            sociedades.append(row["Sociedad2"].strip())

        configuraciones.append({
            "nombre": nombre,
            "cuit": cuit,
            "password": password,
            "descargar_nombre": row.get("Descargar Nombre", "").strip().upper() == "X",
            "descargar_sociedades": len(sociedades) > 0,
            "sociedades": sociedades
        })

    return configuraciones

# Ruta para testing
config_file_path = (
    "C:\\Users\\Julian\\Desktop\\Programacion\\Proyectos\\MarianoMortero\\V1\\Clientes 2025 juli.xlsm"
)

def contar_filas_utiles(ruta_excel, start_row=9, col_clave="A"):
    """
    Cuenta cuántas filas con datos reales hay a partir de start_row.
    Usa la columna clave (por defecto A) para saber si la fila está vacía.
    """
    wb = load_workbook(ruta_excel, read_only=True, data_only=True)
    ws = wb.active

    fila_actual = start_row
    while True:
        valor = ws[f"{col_clave}{fila_actual}"].value
        if valor is None or str(valor).strip() == "":
            break
        fila_actual += 1

    return fila_actual - start_row  # cantidad de filas con datos

# -------- helpers --------
def _digits11(s: str) -> str:
    """Devuelve solo los dígitos (11) de un CUIT o '' si no cumple."""
    d = re.sub(r"\D", "", str(s))
    return d if len(d) == 11 else ""

def _es_persona(cuit: str) -> bool:
    """Regla pedida: persona = CUIT que NO empieza por 30."""
    d = _digits11(cuit)
    return bool(d) and not d.startswith("30")

# -------- Función 1: indexar sociedades --------
def indexar_sociedades(df):
    """
    Devuelve un diccionario {cuit_dueno: [sociedades]}.
    Sociedad = fila cuyo CUIT empieza por 30 y cuyo 'tipo monotributo' trae el CUIT del dueño.
    """
    sociedades_por_dueno = {}
    for _, r in df.iterrows():
        cuit_fila  = _digits11(r["cuit "])             # CUIT de la fila (sociedad o persona)
        dueno_cuit = _digits11(r["tipo monotributo"])  # CUIT del dueño (si aplica)

        if cuit_fila.startswith("30") and dueno_cuit:
            sociedades_por_dueno.setdefault(dueno_cuit, []).append({
                "nombre": str(r["Clientes"]).strip(),
                "cuit": cuit_fila
            })
    return sociedades_por_dueno

# -------- Función 2: construir config --------
def construir_config(df, sociedades_por_dueno):
    """
    Recorre solo personas y les anida sus sociedades.
    Devuelve una lista con la estructura final.
    """
    config = []
    vistos = set()

    for _, row in df.iterrows():
        nombre   = str(row["Clientes"]).strip()
        password = str(row["Clave Afip"]).strip()
        email    = str(row["e-mail"]).strip()
        cuit_p   = _digits11(row["cuit "])

        if not _es_persona(cuit_p):
            continue
        if cuit_p in vistos:
            continue

        config.append({
            "nombre": nombre,
            "cuit": cuit_p,
            "password": password,
            "email": email,
            "sociedades": sociedades_por_dueno.get(cuit_p, [])
        })
        vistos.add(cuit_p)

    return config

# -------- Función principal --------
def procesar_config(df):
    """
    Procesa el DataFrame y devuelve la lista de configuraciones
    (usuarios con sociedades anidadas).
    """
    sociedades = indexar_sociedades(df)
    return construir_config(df, sociedades)

def configuracion(ruta_excel):
    """
    Lee la configuración desde un Excel, detectando automáticamente la cantidad de filas útiles.
    """
    # Columnas que realmente necesitamos
    columnas_utiles = ["Clientes", "cuit ", "Clave Afip", "e-mail", "tipo monotributo"]

    # Detectar cuántas filas reales hay (datos empiezan en fila 10, col A = "Clientes")
    n_filas = contar_filas_utiles(ruta_excel, start_row=10, col_clave="A")

    # Leer solo esas filas y columnas
    df = pd.read_excel(
        ruta_excel,
        skiprows=8,  # encabezado en fila 9
        nrows=n_filas,
        usecols=columnas_utiles,
        engine="openpyxl"
    ).fillna("")

    # Limpieza de consola y vista previa
    # system('cls')
    print(df.head())

    config = []
    # Procesar fila por fila
    for _, row in df.iterrows():
        nombre = row["Clientes"].strip()
        password = row["Clave Afip"].strip()
        email = row["e-mail"].strip()
        tipo = str(row["tipo monotributo"]).strip()
        if re.fullmatch(r"\b\d{11}\b", tipo):
            cuit = str(row["tipo monotributo"]).strip()
        else:
            cuit = str(row["cuit "]).strip()

        print(f"Cliente: {nombre} | CUIT (usuario): {cuit} | Clave: {password} | Email: {email}")

        config.append({
            "nombre": nombre,
            "cuit": cuit,
            "password": password,
            "email": email
            })
    
    return config

# Ejecutar función
# configuracion(config_file_path)

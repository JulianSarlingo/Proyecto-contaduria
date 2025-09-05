from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait

from selenium.webdriver.chrome.options import Options
import time
from datetime import datetime, timedelta
import calendar
import pandas as pd

from openpyxl import load_workbook

config_file_path = "C:\\Users\\Julian\\Desktop\\Programacion\\Proyectos\\MarianoMortero\\V1"
config_file_path = config_file_path+"\\Clientes 2025 juli.xlsm"

def contar_filas_utiles(ruta, sheet_name=None, start_row=9, col_clave="A"):
    wb = load_workbook(ruta, read_only=True, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active
    
    fila_final = start_row
    for row in ws.iter_rows(min_row=start_row, max_col=1, values_only=True):
        if not any(row):  # fila vacía
            break
        fila_final += 1
    return fila_final - start_row  # cantidad de filas útiles

# Ejemplo de uso
n_filas = contar_filas_utiles(config_file_path, start_row=9)
df = pd.read_excel(config_file_path, skiprows=8, nrows=n_filas, engine="openpyxl").fillna("")








def obtener_fechas_mes_pasado_formato_ddMMAAAA():
    """
    Obtiene el primer y último día del mes pasado en formato ddMMyyyy.

    Retorna:
        tuple: Una tupla que contiene dos cadenas de texto (strings):
               (primer_dia_mes_pasado_str, ultimo_dia_mes_pasado_str)
    """
    hoy = datetime.now()

    # Calculamos el primer día del mes actual.
    # Por ejemplo, si hoy es 2025-07-19, esto será 2025-07-01.
    primer_dia_mes_actual = hoy.replace(day=1)

    # Restamos un día para obtener el último día del mes pasado.
    # Si primer_dia_mes_actual es 2025-07-01, esto será 2025-06-30.
    ultimo_dia_mes_pasado_dt = primer_dia_mes_actual - timedelta(days=1)

    # El primer día del mes pasado es simplemente el último día del mes pasado
    # con el día fijado a 1.
    # Si ultimo_dia_mes_pasado_dt es 2025-06-30, esto será 2025-06-01.
    primer_dia_mes_pasado_dt = ultimo_dia_mes_pasado_dt.replace(day=1)

    # Formateamos las fechas a 'ddMMyyyy'
    primer_dia_str = primer_dia_mes_pasado_dt.strftime("%d%m%Y")
    ultimo_dia_str = ultimo_dia_mes_pasado_dt.strftime("%d%m%Y")

    return primer_dia_str, ultimo_dia_str

# ruta_origen = "C:\\Users\\Julian\\Desktop\\Programacion\\Proyectos\\MarianoMortero\\Datos Afip\\"
# ruta_descargas = "C:\\Users\\Julian\\Downloads"
# Inicializador de datos

def init():
    config_file_path = "C:\\Users\\Julian\\Desktop\\Programacion\\Proyectos\\MarianoMortero\\V1"
    config_file_path = config_file_path+"\\Clientes 2025 juli.xlsm"
    # config_file_path = ruta_base+"Clientes 2025 juli.xlsm"

    df = pd.read_excel(config_file_path, skiprows=8, engine="openpyxl").fillna("")

    configuraciones = []

    print(df)

# Ejemplo de uso:
if __name__ == "__main__":
    # primer_dia, ultimo_dia = obtener_fechas_mes_pasado_formato_ddMMAAAA()
    # print(f"El primer día del mes pasado fue: {primer_dia}")
    # print(f"El último día del mes pasado fue: {ultimo_dia}")

    init()